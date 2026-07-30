"""
setup_data.py — 資料處理工具
A  開始編輯：JSON → Excel → 開啟檔案
B  完成編輯：Excel → JSON → 正規化 → 寫回 Excel
0  進階單步執行
"""
import csv
import json
import os
import random
import re
import shutil
import subprocess
import sys
import time
import warnings
import xml.etree.ElementTree as ET

warnings.filterwarnings('ignore')

def install(pkg):
    subprocess.check_call([sys.executable, '-m', 'pip', 'install', pkg, '-q'])

try:
    import requests
except ImportError:
    print('安裝 requests 中...')
    install('requests')
    import requests

try:
    import openpyxl
    from openpyxl.styles import PatternFill, Font, Alignment, numbers
    from openpyxl.worksheet.datavalidation import DataValidation
    from openpyxl.utils import get_column_letter
except ImportError:
    print('安裝 openpyxl 中...')
    install('openpyxl')
    import openpyxl
    from openpyxl.styles import PatternFill, Font, Alignment, numbers
    from openpyxl.worksheet.datavalidation import DataValidation
    from openpyxl.utils import get_column_letter

import urllib3
urllib3.disable_warnings(urllib3.exceptions.InsecureRequestWarning)

tools_dir  = os.path.dirname(os.path.abspath(__file__))
root_dir   = os.path.dirname(tools_dir)
json_path  = os.path.join(root_dir, 'data', 'data.json')
xlsx_path  = os.path.join(tools_dir, 'data.xlsx')
dist_path  = os.path.join(root_dir, 'data', 'districts.json')
stamp_path = os.path.join(tools_dir, '.xlsx_stamp')

XLSX_BACKUP_KEEP = 5   # tools/data_xlsx_backup_*.xlsx 保留份數

# ── 共用 I/O ──────────────────────────────────────────────────────────────────
def load_data():
    with open(json_path, 'r', encoding='utf-8') as f:
        return json.load(f)

def save_data(rows):
    with open(json_path, 'w', encoding='utf-8') as f:
        json.dump(rows, f, ensure_ascii=False, indent=2)

# ── data.xlsx 防覆寫：stamp 偵測 + 覆寫前備份 ─────────────────────────────────
def _write_xlsx_stamp():
    """記下本工具產生 data.xlsx 的 mtime，供 B 路徑判斷有無未回寫的編輯"""
    try:
        with open(stamp_path, 'w', encoding='utf-8') as f:
            f.write(str(os.path.getmtime(xlsx_path)))
    except Exception as e:
        print(f'  ⚠  stamp 寫入失敗（不影響資料）：{e}')

def _xlsx_edited_since_generated():
    """data.xlsx 是否在本工具產生後被改過（= 有未回寫的 Excel 編輯）

    沒有 stamp 或讀不到時一律回傳 True：寧可多問一次，不要默默蓋掉。
    """
    if not os.path.exists(xlsx_path):
        return False
    if not os.path.exists(stamp_path):
        return True
    try:
        with open(stamp_path, 'r', encoding='utf-8') as f:
            stamped = float(f.read().strip())
    except Exception:
        return True
    # 容忍 1 秒檔案系統誤差；人工編輯不可能落在產生後 1 秒內
    return abs(os.path.getmtime(xlsx_path) - stamped) > 1

def _backup_xlsx():
    """覆寫前無條件備份 data.xlsx → tools/data_xlsx_backup_<ts>.xlsx（留最近 N 份）"""
    if not os.path.exists(xlsx_path):
        return None
    import datetime as _dt3
    import glob as _glob2
    ts     = _dt3.datetime.now().strftime('%Y%m%d_%H%M%S')
    target = os.path.join(tools_dir, f'data_xlsx_backup_{ts}.xlsx')
    try:
        shutil.copy2(xlsx_path, target)
    except Exception as e:
        print(f'  ⚠  data.xlsx 備份失敗：{e}')
        return None
    print(f'  💾 已備份舊 Excel：tools/{os.path.basename(target)}')
    # 檔名時間戳為固定寬度，字串排序即時間排序
    olds = sorted(_glob2.glob(os.path.join(tools_dir, 'data_xlsx_backup_*.xlsx')))
    for old in olds[:-XLSX_BACKUP_KEEP]:
        try:
            os.remove(old)
        except Exception:
            pass
    return target

def load_districts():
    if not os.path.exists(dist_path):
        return {}
    with open(dist_path, 'r', encoding='utf-8') as f:
        return json.load(f)

def section(num, title):
    print()
    print('─' * 54)
    print(f'  [{num}]  {title}')
    print('─' * 54)

# ════════════════════════════════════════════════════════════════════════════════
# STEP 1：更新行政區清單
# ════════════════════════════════════════════════════════════════════════════════
def step_update_districts():
    section(1, '更新行政區清單（內政部 API）')

    API_BASE = 'https://api.nlsc.gov.tw/other/ListTown1'
    districts_raw = {}

    for code in 'ABCDEFGHIJKLMNOPQRSTUVWXYZ':
        try:
            r = requests.get(f'{API_BASE}/{code}', timeout=10, verify=False)
            if r.status_code != 200 or '<townItem>' not in r.text:
                continue
            root  = ET.fromstring(r.content)
            items = root.findall('townItem')
            if items:
                districts_raw[code] = [item.findtext('townname') for item in items]
                print(f'    {code}: {len(districts_raw[code])} 個鄉鎮市區')
        except Exception as e:
            print(f'    {code}: 失敗 ({e})')

    county_names = {}
    try:
        r = requests.get('https://api.nlsc.gov.tw/other/ListCounty', timeout=10, verify=False)
        root = ET.fromstring(r.content)
        for item in root.findall('countyItem'):
            county_names[item.findtext('countycode')] = item.findtext('countyname')
    except Exception as e:
        print(f'    縣市名稱 API 失敗: {e}')

    districts = {county_names.get(c, c): towns for c, towns in districts_raw.items()}
    with open(dist_path, 'w', encoding='utf-8') as f:
        json.dump(districts, f, ensure_ascii=False, indent=2)

    total = sum(len(v) for v in districts.values())
    print(f'\n  ✅ 完成：{len(districts)} 縣市，{total} 鄉鎮市區')
    return len(districts)

# ════════════════════════════════════════════════════════════════════════════════
# STEP 2：補縣市／鄉鎮市區
# ════════════════════════════════════════════════════════════════════════════════
def step_fill_city_district():
    section(2, '補縣市／鄉鎮市區')

    districts = load_districts()
    if not districts:
        print('  ⚠  找不到 districts.json，請先執行步驟 1')
        return 0

    rows = load_data()

    def parse(addr):
        if not addr:
            return '', ''
        s = re.sub(r'^\d{3,6}', '', addr.replace('台', '臺')).strip()
        for county, towns in districts.items():
            if s.startswith(county):
                rest = s[len(county):]
                for town in towns:
                    if rest.startswith(town):
                        return county, town
        return '', ''

    updated = 0
    for row in rows:
        addr = row.get('地址', '')
        if addr:
            cleaned = re.sub(r'^\d{3,6}', '', addr).strip()
            if cleaned != addr:
                row['地址'] = cleaned

        # 永遠重新從地址解析，若有新值則覆蓋（處理行政區升格等名稱異動）
        city, dist = parse(row.get('地址', ''))
        changed = False
        if city and row.get('縣市') != city:
            print(f'    ✓ {row["店名"]}：縣市 {row.get("縣市","（空）")!r} → {city!r}')
            row['縣市'] = city
            changed = True
        if dist and row.get('鄉鎮市區') != dist:
            print(f'    ✓ {row["店名"]}：鄉鎮市區 {row.get("鄉鎮市區","（空）")!r} → {dist!r}')
            row['鄉鎮市區'] = dist
            changed = True
        if changed:
            updated += 1
        elif not city:
            print(f'    ✗ {row["店名"]}：地址解析失敗')

    save_data(rows)
    print(f'\n  ✅ 完成：補填 {updated} 筆（共 {len(rows)} 筆）')
    return updated

# ════════════════════════════════════════════════════════════════════════════════
# STEP 3：補 lat/lng 座標
# ════════════════════════════════════════════════════════════════════════════════
def step_geocode():
    section(3, '補 lat/lng 座標')

    print('  模式選擇（直接 Enter = 只補缺少座標）：')
    print('    1. 只補缺少座標的店家  ← 預設')
    print('    2. 重新更正所有有 Map URL 的店家（修正舊座標精度）')
    mode = input('  請輸入 1 或 2：').strip() or '1'

    rows  = load_data()
    total = len(rows)
    if mode == '2':
        to_geocode = [r for r in rows if r.get('Map', '').startswith('http') or not r.get('lat')]
        print(f'  重新 geocode：{len(to_geocode)} 筆（共 {total} 筆）')
    else:
        to_geocode = [r for r in rows if not r.get('lat') or not r.get('lng')]
        print(f'  需要 geocode：{len(to_geocode)} 筆（共 {total} 筆）')

    if not to_geocode:
        print('  ✅ 無需處理')
        return 0

    UA = {'User-Agent': 'Mozilla/5.0 (Windows NT 10.0; Win64; x64) AppleWebKit/537.36'}
    updated    = 0
    failed     = []
    consecutive = 0
    MAX_CONSEC  = 5

    def from_map_url(url):
        if not url or not url.startswith('http'):
            return None, None
        r = requests.get(url, headers=UA, timeout=10, verify=False, allow_redirects=True)
        # !3d!4d 是 Google Maps 標記點的精確座標
        # 不使用 /@ 的 fallback：那是地圖視角中心，縮放狀態不同會漂移到海上
        m = re.search(r'!3d(-?\d+\.\d+)!4d(-?\d+\.\d+)', r.url)
        if m:
            return float(m.group(1)), float(m.group(2))
        return None, None

    def from_nominatim(address):
        r = requests.get('https://nominatim.openstreetmap.org/search',
            params={'q': address, 'format': 'json', 'limit': 1},
            headers=UA, timeout=10, verify=False)
        res = r.json()
        return (float(res[0]['lat']), float(res[0]['lon'])) if res else (None, None)

    def in_taiwan(lat, lng):
        """台灣合理座標範圍（含金門 ~118.3E、馬祖 ~119.9E、蘭嶼、綠島）"""
        return 21.5 <= lat <= 26.5 and 118.0 <= lng <= 122.5

    for i, row in enumerate(to_geocode):
        name    = row.get('店名', '')
        address = row.get('地址', '') or name
        print(f'  [{i+1}/{len(to_geocode)}] {name}')
        try:
            lat, lng = from_map_url(row.get('Map', ''))
            if lat and in_taiwan(lat, lng):
                print(f'    ✓ (Map URL) {lat:.6f}, {lng:.6f}')
            else:
                if lat:
                    print(f'    ✗ (Map URL) 座標超出台灣範圍：{lat:.6f}, {lng:.6f}，改用 Nominatim')
                    lat = lng = None
                lat, lng = from_nominatim(address)
                if lat and in_taiwan(lat, lng):
                    print(f'    ✓ (Nominatim) {lat:.6f}, {lng:.6f}')
                elif lat:
                    print(f'    ✗ (Nominatim) 座標超出台灣範圍：{lat:.6f}, {lng:.6f}')
                    lat = lng = None

            if lat:
                row['lat'] = lat
                row['lng'] = lng
                updated   += 1
                consecutive = 0
            else:
                failed.append(name)
                consecutive += 1
                print(f'    ✗ 找不到座標（連續失敗 {consecutive}/{MAX_CONSEC}）')
        except Exception as e:
            failed.append(name)
            consecutive += 1
            print(f'    ✗ 錯誤：{e}')

        if consecutive >= MAX_CONSEC:
            print(f'\n  ⚠  連續失敗 {MAX_CONSEC} 筆，中斷作業')
            break
        time.sleep(1.1)

    save_data(rows)
    print(f'\n  ✅ 完成：更新 {updated} 筆（共 {total} 筆）')
    if failed:
        print(f'  ⚠  無法取得座標：{", ".join(failed)}')
    return updated

# ════════════════════════════════════════════════════════════════════════════════
# STEP 4：正規化營業時段
# ════════════════════════════════════════════════════════════════════════════════
HOURS_FIELDS = ['週一', '週二', '週三', '週四', '週五', '週六', '週日', '營業時段']

def normalize_hours(value):
    if not isinstance(value, str) or not value.strip():
        return value
    v = value.strip()
    v = re.sub(r'(?<=\d)[—\-~～](?=\d)', '–', v)
    segments = re.findall(r'\d{1,2}:\d{2}–\d{1,2}:\d{2}', v)
    return '、'.join(segments) if segments else v

def step_normalize_hours():
    section(4, '正規化營業時段格式')

    rows    = load_data()
    updated = 0
    for row in rows:
        for field in HOURS_FIELDS:
            original   = row.get(field, '')
            normalized = normalize_hours(original)
            if normalized != original:
                row[field] = normalized
                updated   += 1
                print(f'    {row["店名"]} [{field}]  {original!r} → {normalized!r}')

    save_data(rows)
    print(f'\n  ✅ 完成：更新 {updated} 個欄位（共 {len(rows)} 筆）')
    return updated

# ════════════════════════════════════════════════════════════════════════════════
# STEP 5：正規化星期排序
# ════════════════════════════════════════════════════════════════════════════════
DAY_ORDER  = {'一': 1, '二': 2, '三': 3, '四': 4, '五': 5, '六': 6, '日': 7}
DAY_FIELDS = ['營業日', '店休日']

def normalize_days(value):
    if not value or not isinstance(value, str):
        return value
    parts = [p.strip() for p in value.split(',')]
    if not all(p in DAY_ORDER for p in parts if p):
        return value
    return ', '.join(sorted(parts, key=lambda d: DAY_ORDER.get(d, 99)))

def step_normalize_days():
    section(5, '正規化星期排序')

    rows    = load_data()
    updated = 0
    for row in rows:
        for field in DAY_FIELDS:
            original   = row.get(field, '')
            normalized = normalize_days(original)
            if normalized != original:
                row[field] = normalized
                updated   += 1
                print(f'    {row["店名"]} [{field}]  {original!r} → {normalized!r}')

    save_data(rows)
    print(f'\n  ✅ 完成：更新 {updated} 個欄位（共 {len(rows)} 筆）')
    return updated

# ════════════════════════════════════════════════════════════════════════════════
# STEP 6：正規化開幕日期
# ════════════════════════════════════════════════════════════════════════════════
DATE_FIELDS = ['開幕日', '歇業日']

def normalize_date(value):
    """各種日期格式統一為 YYYY-MM-DD，無法辨識則原樣返回"""
    if not value or not isinstance(value, str) or not value.strip():
        return value
    v = value.strip()

    # 已經是標準格式
    if re.match(r'^\d{4}-\d{2}-\d{2}$', v):
        return v

    # YYYY-MM-DD HH:MM:SS 或 YYYY-MM-DD HH:MM（Excel datetime）
    m = re.match(r'^(\d{4})-(\d{1,2})-(\d{1,2})[\sT]\d{1,2}:\d{2}', v)
    if m:
        return f'{m.group(1)}-{int(m.group(2)):02d}-{int(m.group(3)):02d}'

    # YYYY/MM/DD 或 YYYY/M/D
    m = re.match(r'^(\d{4})/(\d{1,2})/(\d{1,2})$', v)
    if m:
        return f'{m.group(1)}-{int(m.group(2)):02d}-{int(m.group(3)):02d}'

    # YYYY.MM.DD 或 YYYY.M.D
    m = re.match(r'^(\d{4})\.(\d{1,2})\.(\d{1,2})$', v)
    if m:
        return f'{m.group(1)}-{int(m.group(2)):02d}-{int(m.group(3)):02d}'

    # YYYY-M-D（有破折號但未補零）
    m = re.match(r'^(\d{4})-(\d{1,2})-(\d{1,2})$', v)
    if m:
        return f'{m.group(1)}-{int(m.group(2)):02d}-{int(m.group(3)):02d}'

    # YYYYMMDD（純數字 8 碼）
    m = re.match(r'^(\d{4})(\d{2})(\d{2})$', v)
    if m:
        return f'{m.group(1)}-{m.group(2)}-{m.group(3)}'

    return v  # 無法辨識，原樣返回

# 月精度開幕日（只知年月、不知日）→ 正規化為 YYYY-MM 並推導開幕月份。
# 只套用在「開幕日」；歇業日維持只認完整日期（自動歇業判定需要日）。
# 注意：以下四種寫法 normalize_date() 都會原樣返回，因此不會誤入 failed 清單。
MONTH_ONLY_PATTERNS = (
    r'^(\d{4})-(\d{1,2})$',      # 2014-07 / 2014-7
    r'^(\d{4})/(\d{1,2})$',      # 2014/07 / 2014/7
    r'^(\d{4})\.(\d{1,2})$',     # 2014.7
    r'^(\d{4})年(\d{1,2})月$',   # 2014年7月
)

def parse_month_only(value):
    """回傳 ('YYYY-MM', month) 或 None（無法辨識／月份不在 1-12）"""
    if not value or not isinstance(value, str):
        return None
    v = value.strip()
    for pat in MONTH_ONLY_PATTERNS:
        m = re.match(pat, v)
        if m:
            mo = int(m.group(2))
            return (f'{m.group(1)}-{mo:02d}', mo) if 1 <= mo <= 12 else None
    return None

def step_normalize_dates():
    section(6, '正規化開幕日 / 歇業日（→ YYYY-MM-DD）')

    rows    = load_data()
    updated = 0
    failed  = []

    for row in rows:
        for field in DATE_FIELDS:
            original = row.get(field, '')
            if not original:
                continue
            normalized = normalize_date(original)
            if normalized == original:
                continue
            if re.match(r'^\d{4}-\d{2}-\d{2}$', normalized):
                row[field] = normalized
                updated   += 1
            else:
                failed.append((row['店名'], field, original))

        # 開幕日正規化後同步更新開幕月份
        d = str(row.get('開幕日', '')).strip()
        if re.match(r'^\d{4}-\d{2}-\d{2}$', d):
            month = int(d.split('-')[1])
            if row.get('開幕月份') != month:
                row['開幕月份'] = month
                updated += 1
        else:
            parsed = parse_month_only(d)
            if parsed:
                norm, month = parsed
                if d != norm:
                    row['開幕日'] = norm
                    updated += 1
                if row.get('開幕月份') != month:
                    row['開幕月份'] = month
                    updated += 1

    save_data(rows)
    print(f'\n  ✅ 完成：更新 {updated} 個欄位（共 {len(rows)} 筆）')
    if failed:
        print(f'  ⚠  無法辨識格式（請手動修正）：')
        for name, field, val in failed:
            print(f'      {name} [{field}] = {val!r}')
    return updated

# ════════════════════════════════════════════════════════════════════════════════
# STEP 7：分配店家 ID
# ════════════════════════════════════════════════════════════════════════════════
CITY_CODE = {
    '臺北市': 'A', '台北市': 'A',
    '新北市': 'B',
    '桃園市': 'C',
    '臺中市': 'D', '台中市': 'D',
    '臺南市': 'E', '台南市': 'E',
    '高雄市': 'F',
    '基隆市': 'G',
    '新竹市': 'H',
    '新竹縣': 'I',
    '苗栗縣': 'J',
    '彰化縣': 'K',
    '南投縣': 'L',
    '雲林縣': 'M',
    '嘉義市': 'N',
    '嘉義縣': 'O',
    '屏東縣': 'P',
    '宜蘭縣': 'Q',
    '花蓮縣': 'R',
    '臺東縣': 'S', '台東縣': 'S',
    '澎湖縣': 'T',
    '金門縣': 'U',
    '連江縣': 'V',
}
ID_RE = re.compile(r'^[A-Z]\d{5}$')

CODE_TO_CITY = {
    'A': '臺北市', 'B': '新北市', 'C': '桃園市', 'D': '臺中市',
    'E': '臺南市', 'F': '高雄市', 'G': '基隆市', 'H': '新竹市',
    'I': '新竹縣', 'J': '苗栗縣', 'K': '彰化縣', 'L': '南投縣',
    'M': '雲林縣', 'N': '嘉義市', 'O': '嘉義縣', 'P': '屏東縣',
    'Q': '宜蘭縣', 'R': '花蓮縣', 'S': '臺東縣', 'T': '澎湖縣',
    'U': '金門縣', 'V': '連江縣', 'Z': '未知縣市',
}

def _get_city_for_id(row):
    city = str(row.get('縣市', '')).strip()
    if not city:
        addr = str(row.get('地址', '')).strip()
        addr = re.sub(r'^\d{3,6}', '', addr)
        city = addr[:3]
    return city.replace('台', '臺')

counters_path = os.path.join(root_dir, 'data', 'id_counters.json')

def _load_counters():
    if os.path.exists(counters_path):
        with open(counters_path, 'r', encoding='utf-8') as f:
            return json.load(f)
    return {}

def _save_counters(city_max):
    with open(counters_path, 'w', encoding='utf-8') as f:
        json.dump(city_max, f, ensure_ascii=False, indent=2, sort_keys=True)

def step_assign_ids():
    section(7, '分配店家 ID（縣市代碼 + 5位流水號）')

    rows = load_data()

    # 從 id_counters.json 載入歷史最大值（防止因資料列被誤刪導致 ID 重用）
    city_max = _load_counters()
    print(f'  📋 讀取歷史計數器：{dict(sorted(city_max.items()))}')

    # 再與現有資料的 ID 取最大值（兩者都納入，只增不減）
    for row in rows:
        eid = str(row.get('ID', '')).strip()
        if ID_RE.match(eid):
            letter = eid[0]
            city_max[letter] = max(city_max.get(letter, 0), int(eid[1:]))

    assigned = 0
    for row in rows:
        eid = str(row.get('ID', '')).strip()
        if ID_RE.match(eid):
            continue
        city   = _get_city_for_id(row)
        letter = CITY_CODE.get(city, 'Z')
        next_n = city_max.get(letter, 0) + 1
        city_max[letter] = next_n
        row['ID'] = f'{letter}{next_n:05d}'
        assigned += 1
        print(f'    ✓ {row["店名"]}  →  {row["ID"]}')

    # 確保 ID 排在第一欄
    rows = [{'ID': r.get('ID', ''), **{k: v for k, v in r.items() if k != 'ID'}} for r in rows]

    save_data(rows)

    # 寫回計數器（只增不減，是唯一的安全防線）
    _save_counters(city_max)
    print(f'  💾 計數器已更新 → data/id_counters.json')

    print(f'\n  ✅ 完成：新分配 {assigned} 筆，共 {len(rows)} 筆')

    city_counts = {}
    for row in rows:
        letter = row.get('ID', 'Z')[0] if row.get('ID') else 'Z'
        city_counts[letter] = city_counts.get(letter, 0) + 1
    print('\n  各縣市店家數量：')
    for letter in sorted(city_counts):
        print(f'    {letter} {CODE_TO_CITY.get(letter, letter)}: {city_counts[letter]} 間')

    return assigned

# ════════════════════════════════════════════════════════════════════════════════
# Excel ↔ JSON 轉換
# ════════════════════════════════════════════════════════════════════════════════
DATE_TEXT_FIELDS = {'開幕日', 'ID'}

def step_excel_to_json():
    section('E', 'Excel → JSON')
    if not os.path.exists(xlsx_path):
        print('  ❌ 找不到 data.xlsx，請先執行 A【開始編輯】')
        return False
    print('  📂 讀取 data.xlsx...')
    wb = openpyxl.load_workbook(xlsx_path)
    ws = wb.active
    rows_raw = list(ws.values)
    if not rows_raw:
        print('  ❌ Excel 是空的')
        return False
    headers = [str(h).strip() for h in rows_raw[0]]
    rows = []
    for row in rows_raw[1:]:
        if all((v is None or str(v).strip() == '') for v in row):
            continue
        obj = {}
        for i, h in enumerate(headers):
            val = row[i] if i < len(row) else ''
            if val is None:
                val = ''
            else:
                val = str(val).strip()
                if val.endswith('.0') and val[:-2].lstrip('-').isdigit():
                    val = val[:-2]
            obj[h] = val
        if obj.get('店名', '').strip():
            rows.append(obj)

    # ── ID 安全保護：用舊 data.json 補回被 Excel 清空的 ID ──────────────
    # 比對 key：(店名, 地址)，兩者都空才視為無法對應
    existing_rows = load_data()
    old_id_map = {}
    for r in existing_rows:
        key = (str(r.get('店名', '')).strip(), str(r.get('地址', '')).strip())
        eid = str(r.get('ID', '')).strip()
        if ID_RE.match(eid):
            old_id_map[key] = eid

    restored = 0
    for r in rows:
        if ID_RE.match(str(r.get('ID', '')).strip()):
            continue  # Excel 已有合法 ID，保留不動
        key = (str(r.get('店名', '')).strip(), str(r.get('地址', '')).strip())
        if key in old_id_map:
            r['ID'] = old_id_map[key]
            restored += 1

    if restored:
        print(f'  🔒 ID 保護：從舊資料補回 {restored} 筆 ID（Excel 中為空）')
    # ────────────────────────────────────────────────────────────────────

    save_data(rows)
    print(f'  ✅ 完成：data.json 已更新（共 {len(rows)} 筆）')
    return True

def step_json_to_excel():
    section('X', 'JSON → Excel（含樣式與下拉驗證）')
    rows = load_data()
    if not rows:
        print('  ❌ data.json 是空的')
        return False
    _backup_xlsx()   # 任何路徑覆寫 data.xlsx 前都先留一份
    print(f'  📝 寫入 {len(rows)} 筆資料...')
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = '店家資料'
    headers = list(rows[0].keys())
    header_fill = PatternFill(start_color='C8272D', end_color='C8272D', fill_type='solid')
    header_font = Font(color='FFFFFF', bold=True, size=11)
    for col_idx, h in enumerate(headers, 1):
        cell = ws.cell(row=1, column=col_idx, value=h)
        cell.fill      = header_fill
        cell.font      = header_font
        cell.alignment = Alignment(horizontal='center', vertical='center')
    for row_idx, row in enumerate(rows, 2):
        for col_idx, h in enumerate(headers, 1):
            cell = ws.cell(row=row_idx, column=col_idx, value=row.get(h, ''))
            if h in DATE_TEXT_FIELDS:
                cell.number_format = numbers.FORMAT_TEXT
    ws.freeze_panes = 'A2'
    for col in ws.columns:
        max_len = max((len(str(cell.value or '')) for cell in col), default=0)
        ws.column_dimensions[col[0].column_letter].width = min(max_len + 2, 40)
    csv_path = os.path.join(tools_dir, 'item_detail.csv')
    if os.path.exists(csv_path):
        with open(csv_path, 'r', encoding='utf-8-sig') as f:
            reader      = csv.reader(f)
            vld_headers = next(reader)
            vld_cols    = {h: [] for h in vld_headers}
            for r in reader:
                for i, h in enumerate(vld_headers):
                    val = r[i].strip() if i < len(r) else ''
                    if val:
                        vld_cols[h].append(val)
        ws_vld = wb.create_sheet('驗證清單')
        ws_vld.sheet_state = 'hidden'
        for col_idx, h in enumerate(vld_headers, 1):
            ws_vld.cell(row=1, column=col_idx, value=h)
            for row_idx, val in enumerate(vld_cols[h], 2):
                ws_vld.cell(row=row_idx, column=col_idx, value=val)
        last_row   = len(rows) + 1
        col_letter = {}
        for col_idx, h in enumerate(headers, 1):
            if h in vld_cols:
                col_letter[h] = get_column_letter(col_idx)
        for vld_col_idx, h in enumerate(vld_headers, 1):
            letter = col_letter.get(h)
            if not letter or not vld_cols[h]:
                continue
            vld_col_letter = get_column_letter(vld_col_idx)
            vld_range = f'驗證清單!${vld_col_letter}$2:${vld_col_letter}${len(vld_cols[h]) + 1}'
            dv = DataValidation(type='list', formula1=vld_range, showDropDown=False, allow_blank=True)
            dv.sqref = f'{letter}2:{letter}{last_row}'
            ws.add_data_validation(dv)
        print('  ✅ 已套用下拉選單驗證')
    wb.save(xlsx_path)
    _write_xlsx_stamp()
    print(f'  ✅ 完成！data.xlsx 已產生（共 {len(rows)} 筆）')
    print(f'  📍 {xlsx_path}')
    return True

# ════════════════════════════════════════════════════════════════════════════════
# STEP 10 / 11：Map 連結標準化
# ════════════════════════════════════════════════════════════════════════════════
def step_normalize_map_urls(mode='new_only'):
    """
    mode='new_only' : 只處理仍是 maps.app.goo.gl 的短連結（C 流程自動呼叫）
    mode='all'      : 重新處理所有有 Map 值的店家
    """
    if mode == 'new_only':
        section(10, 'Map 連結標準化（僅短連結）')
    else:
        section(11, 'Map 連結全部重新掃描')

    rows = load_data()

    if mode == 'new_only':
        targets = [r for r in rows if 'maps.app.goo.gl' in r.get('Map', '')]
    else:
        targets = [r for r in rows if r.get('Map', '').startswith('http')]

    total = len(targets)
    if total == 0:
        print('  ✅ 無需處理')
        return 0

    avg_sec = 6  # (2+10)/2
    est_min = total * avg_sec // 60
    est_sec = total * avg_sec % 60
    print(f'  待處理：{total} 筆，間隔 2～10 秒隨機，預計約 {est_min} 分 {est_sec} 秒')
    if total > 20:
        confirm = input('  確認開始？(Enter / y 繼續，其他取消)：').strip().lower()
        if confirm not in ('', 'y'):
            print('  已取消')
            return 0

    UA  = {'User-Agent': 'Mozilla/5.0 (Windows NT 10.0; Win64; x64) AppleWebKit/537.36'}
    ok  = 0
    fail = 0

    for i, row in enumerate(targets):
        name = row.get('店名', '')
        url  = row.get('Map', '')
        lat  = row.get('lat')
        lng  = row.get('lng')

        try:
            # 1. 展開短連結
            if 'maps.app.goo.gl' in url:
                r   = requests.head(url, headers=UA, allow_redirects=True, timeout=10, verify=False)
                url = r.url

            # 2. 移除 query string（g_ep / skid 等動態參數）
            url = url.split('?')[0]

            # 3. 修正 /@lat,lng,Xz/ → 使用 data.json 座標 + 固定 17z
            if re.search(r'/@[-\d.]+,[-\d.]+,[\d.]+z/', url):
                if lat and lng:
                    url = re.sub(r'/@[-\d.]+,[-\d.]+,[\d.]+z/', f'/@{lat},{lng},17z/', url)
                else:
                    url = re.sub(r'(/@[-\d.]+,[-\d.]+,)[\d.]+z/', r'\g<1>17z/', url)

            row['Map'] = url
            print(f'  [{i+1}/{total}] ✅ {name}')
            ok += 1

        except Exception as e:
            print(f'  [{i+1}/{total}] ❌ {name}：{e}')
            fail += 1

        if i < total - 1:
            delay = random.uniform(2, 10)
            time.sleep(delay)

    save_data(rows)
    print(f'\n  ✅ 完成：成功 {ok} 筆 / 失敗 {fail} 筆')
    return ok

def step_normalize_map_new():
    step_normalize_map_urls(mode='new_only')

def step_normalize_map_all():
    step_normalize_map_urls(mode='all')

# ════════════════════════════════════════════════════════════════════════════════
# STEP 8：依縣市排序
# ════════════════════════════════════════════════════════════════════════════════
CITY_ORDER = [
    # 六都
    '臺北市', '新北市', '桃園市', '臺中市', '臺南市', '高雄市',
    # 其他縣市由北到南
    '基隆市', '新竹市', '新竹縣', '苗栗縣',
    '彰化縣', '南投縣', '雲林縣', '嘉義市', '嘉義縣',
    '屏東縣', '宜蘭縣', '花蓮縣', '臺東縣',
    '澎湖縣', '金門縣', '連江縣',
]
CITY_RANK = {city: i for i, city in enumerate(CITY_ORDER)}

# 純地理由北到南（單步選用）
CITY_ORDER_GEO = [
    '基隆市', '臺北市', '新北市', '桃園市',
    '新竹市', '新竹縣', '苗栗縣', '臺中市',
    '彰化縣', '南投縣', '雲林縣', '嘉義市', '嘉義縣',
    '臺南市', '高雄市', '屏東縣',
    '宜蘭縣', '花蓮縣', '臺東縣',
    '澎湖縣', '金門縣', '連江縣',
]
CITY_RANK_GEO = {city: i for i, city in enumerate(CITY_ORDER_GEO)}

def step_sort(mode='priority'):
    """
    mode='priority' : 六都優先，再由北到南（C 路徑預設）
    mode='geo'      : 純地理由北到南（單步選用）
    """
    if mode == 'geo':
        section(8, '依縣市排序（純地理由北到南）')
        rank = CITY_RANK_GEO
    else:
        section(8, '依縣市排序（六都優先，再由北到南）')
        rank = CITY_RANK

    rows = load_data()

    def sort_key(row):
        city = str(row.get('縣市', '')).strip().replace('台', '臺')
        dist = str(row.get('鄉鎮市區', '')).strip()
        addr = str(row.get('地址', '')).strip()
        return (rank.get(city, 99), city, dist, addr)

    rows.sort(key=sort_key)
    save_data(rows)
    print(f'  ✅ 完成：已排序 {len(rows)} 筆')


def step_sort_interactive():
    """單步選單用：讓使用者選擇排序模式"""
    print('\n  排序模式：')
    print('    1. 六都優先，再由北到南（預設）')
    print('    2. 純地理由北到南（忽略六都優先）')
    m = input('  請選擇 1 或 2（直接 Enter = 1）：').strip() or '1'
    step_sort(mode='geo' if m == '2' else 'priority')

# ════════════════════════════════════════════════════════════════════════════════
# STEP 9：自動更新歇業狀態
# ════════════════════════════════════════════════════════════════════════════════
import datetime as _dt

def step_auto_close():
    section(9, '自動更新歇業狀態（歇業日已過 → 已歇業）')
    rows = load_data()
    today = _dt.date.today().isoformat()   # YYYY-MM-DD
    updated = 0

    for row in rows:
        closing = str(row.get('歇業日', '')).strip()
        if not re.match(r'^\d{4}-\d{2}-\d{2}$', closing):
            continue
        if row.get('營業狀態') != '營業中':
            continue
        if today > closing:
            old = row.get('營業狀態', '（空）')
            print(f'    ✓ {row["店名"]}：歇業日 {closing} 已過，{old!r} → "已歇業"')
            row['營業狀態'] = '已歇業'
            updated += 1

    save_data(rows)
    print(f'\n  ✅ 完成：更新 {updated} 筆')
    return updated

# ════════════════════════════════════════════════════════════════════════════════
# 選單
# ════════════════════════════════════════════════════════════════════════════════
STEPS = [
    (1, '更新行政區清單（內政部 API）',          step_update_districts),
    (2, '補縣市／鄉鎮市區',                      step_fill_city_district),
    (3, '補 lat/lng 座標',                       step_geocode),
    (4, '正規化營業時段格式',                    step_normalize_hours),
    (5, '正規化星期排序',                        step_normalize_days),
    (6, '正規化開幕日 / 歇業日（→ YYYY-MM-DD）', step_normalize_dates),
    (7, '分配店家 ID',                           step_assign_ids),
    (8,  '依縣市排序',                             step_sort_interactive),
    (9,  '自動更新歇業狀態',                      step_auto_close),
    (10, 'Map 連結標準化（僅短連結）',             step_normalize_map_new),
    (11, 'Map 連結全部重新掃描',                  step_normalize_map_all),
]

def show_menu():
    print()
    print('╔' + '═' * 52 + '╗')
    print('║{:^52}║'.format('資料處理工具　Setup Data'))
    print('╠' + '═' * 52 + '╣')
    print('║  A  【拉取最新】git pull{:<28}║'.format(''))
    print('║  B  【開始編輯】JSON → Excel，開啟檔案{:<12}║'.format(''))
    print('║  C  【完成編輯】正規化 → Excel（無xlsx直接跑）{:<5}║'.format(''))
    print('║  D  【推上遠端】git push data.json + 計數器{:<8}║'.format(''))
    print('║  ' + '─' * 49 + '║')
    print('║  0  進階單步執行（含 10 Map標準化 / 11 全掃描）{:<4}║'.format(''))
    print('║  ' + '─' * 49 + '║')
    print('║  q  離開{:<43}║'.format(''))
    print('╚' + '═' * 52 + '╝')

def show_advanced_menu():
    print()
    print('╔' + '═' * 52 + '╗')
    print('║{:^52}║'.format('進階單步執行'))
    print('╠' + '═' * 52 + '╣')
    for num, desc, _ in STEPS:
        print(f'║  {num}  {desc:<46}║')
    print('║  ' + '─' * 49 + '║')
    print('║  b  返回主選單{:<37}║'.format(''))
    print('╚' + '═' * 52 + '╝')

def run_path_a():
    print('\n▶ A【拉取最新】git pull')
    result = subprocess.run(
        ['git', 'pull'],
        cwd=root_dir, capture_output=True, text=True, encoding='utf-8'
    )
    print(result.stdout.strip() or result.stderr.strip())
    if result.returncode == 0:
        print('  ✅ 完成')
    else:
        print('  ❌ git pull 失敗')

def open_file(path):
    """跨平台開啟檔案"""
    if sys.platform == 'win32':
        subprocess.Popen(['cmd', '/c', 'start', '', path])
    elif sys.platform == 'darwin':
        subprocess.Popen(['open', path])
    else:
        subprocess.Popen(['xdg-open', path])

def run_path_b():
    print('\n▶ B【開始編輯】JSON → Excel → 開啟檔案')

    # 防呆：data.xlsx 有未回寫的編輯就先攔住（B 會用 data.json 整份蓋掉它）
    if _xlsx_edited_since_generated():
        saved_at = time.strftime('%Y-%m-%d %H:%M:%S',
                                 time.localtime(os.path.getmtime(xlsx_path)))
        print()
        print('  ' + '═' * 50)
        print(f'  ⚠  data.xlsx 在產生後被修改過（最後存檔：{saved_at}）')
        print('     B 會用 data.json 整份覆蓋，你在 Excel 裡的編輯會消失。')
        print('     要把編輯寫回 data.json，請改按 C【完成編輯】。')
        print('  ' + '═' * 50)
        ans = input('  仍要覆蓋？輸入 yes 繼續，其他輸入取消：').strip().lower()
        if ans != 'yes':
            print('  ✅ 已取消，data.xlsx 保持原狀。')
            return

    _backup_data()   # 編輯前先備份，C 路徑成功後自動刪除
    ok = step_json_to_excel()
    if ok:
        print('\n  📂 開啟 Excel...')
        open_file(xlsx_path)

def run_path_d():
    print('\n▶ D【推上遠端】git push data.json + id_counters.json')

    result = subprocess.run(
        ['git', 'add', 'data/data.json', 'data/id_counters.json'],
        cwd=root_dir, capture_output=True, text=True, encoding='utf-8'
    )
    if result.returncode != 0:
        print(f'  ❌ git add 失敗：{result.stderr.strip()}')
        return

    status = subprocess.run(
        ['git', 'diff', '--cached', '--stat'],
        cwd=root_dir, capture_output=True, text=True, encoding='utf-8'
    ).stdout.strip()

    if not status:
        print('  ℹ  無變更，不需要 commit')
        return

    print(f'\n  變更內容：\n{status}\n')
    msg = input('  請輸入 commit 訊息（直接 Enter 使用預設）：').strip()
    if not msg:
        msg = f'更新店家資料'

    result = subprocess.run(
        ['git', 'commit', '-m', msg],
        cwd=root_dir, capture_output=True, text=True, encoding='utf-8'
    )
    if result.returncode != 0:
        print(f'  ❌ commit 失敗：{result.stderr.strip()}')
        return
    print(f'  ✅ Committed')

    print('  🚀 git push...')
    result = subprocess.run(
        ['git', 'push'],
        cwd=root_dir, capture_output=True, text=True, encoding='utf-8'
    )
    if result.returncode == 0:
        print('  ✅ Push 完成！')
    else:
        print(f'  ❌ push 失敗：{result.stderr.strip()}')

# ════════════════════════════════════════════════════════════════════════════════
# C 路徑輔助：備份 + ID 驗證
# ════════════════════════════════════════════════════════════════════════════════
def _backup_data():
    """備份 data.json → data/data_backup_YYYYMMDD_HHMMSS.json（B 路徑於開始編輯前呼叫）"""
    import datetime as _dt2
    ts          = _dt2.datetime.now().strftime('%Y%m%d_%H%M%S')
    backup_path = os.path.join(root_dir, 'data', f'data_backup_{ts}.json')
    shutil.copy2(json_path, backup_path)
    print(f'  💾 備份完成：data/data_backup_{ts}.json')
    return backup_path


def _delete_backups():
    """刪除 data/ 下所有 data_backup_*.json（C 路徑成功後呼叫）"""
    import glob as _glob
    pattern = os.path.join(root_dir, 'data', 'data_backup_*.json')
    files   = _glob.glob(pattern)
    if not files:
        return
    for f in files:
        try:
            os.remove(f)
            print(f'  🗑  備份已自動刪除（{os.path.basename(f)}）')
        except Exception as e:
            print(f'  ⚠  備份刪除失敗：{e}')


def step_validate_ids(old_row_count):
    """
    驗證 ID 完整性（在 excel_to_json 之後、assign_ids 之前執行）：
      1. 新舊行數差異
      2. 重複 ID 偵測
      3. 各縣市 ID 序列是否連續（無跳號）
      4. 與 id_counters 比對
    回傳 has_critical (bool) — True 表示發現重複或跳號
    """
    section('V', 'ID 驗證（行數 / 重複 / 序列完整性 / 計數器）')

    rows     = load_data()
    counters = _load_counters()

    new_row_count = len(rows)
    diff          = new_row_count - old_row_count

    # 1. 行數比對
    if diff == 0:
        print(f'  📊 總行數：{new_row_count} 筆（與舊資料相同）')
    elif diff > 0:
        print(f'  📊 總行數：{new_row_count} 筆（較舊資料 +{diff} 筆，有新增）')
    else:
        print(f'  ⚠  總行數：{new_row_count} 筆（較舊資料 {diff} 筆，有資料減少！）')

    # 2. 收集各縣市 ID
    city_ids    = {}    # letter → sorted list of int
    id_to_names = {}    # id_str → [shop_name, ...]

    for row in rows:
        eid = str(row.get('ID', '')).strip()
        if ID_RE.match(eid):
            letter = eid[0]
            num    = int(eid[1:])
            city_ids.setdefault(letter, []).append(num)
            id_to_names.setdefault(eid, []).append(row.get('店名', '（未知）'))

    has_critical = False

    # 3. 重複 ID 偵測
    duplicates = {eid: names for eid, names in id_to_names.items() if len(names) > 1}
    if duplicates:
        has_critical = True
        print(f'\n  ❌ 發現重複 ID（{len(duplicates)} 組）：')
        for eid, names in sorted(duplicates.items()):
            print(f'      {eid}：{" / ".join(names)}')
    else:
        print(f'  ✅ 無重複 ID')

    # 4. 各縣市序列完整性 + id_counters 比對
    print(f'\n  各縣市 ID 驗證：')
    for letter in sorted(city_ids.keys()):
        nums         = sorted(city_ids[letter])
        actual_count = len(nums)
        max_num      = nums[-1]
        counter_val  = counters.get(letter, 0)
        city_name    = CODE_TO_CITY.get(letter, letter)

        # 跳號：期望 1~max_num 的完整集合
        missing         = sorted(set(range(1, max_num + 1)) - set(nums))
        counter_mismatch = counter_val != max_num

        if missing or counter_mismatch:
            has_critical = True
            parts = []
            if missing:
                miss_strs = [f'{letter}{n:05d}' for n in missing[:10]]
                more_note = f'...等共 {len(missing)} 個' if len(missing) > 10 else ''
                parts.append(f'缺號：{", ".join(miss_strs)}{more_note}')
            if counter_mismatch:
                parts.append(f'計數器={counter_val} 與最大 ID {letter}{max_num:05d} 不符')
            print(f'  ❌ {letter} {city_name}：{actual_count} 筆，'
                  f'最大={letter}{max_num:05d}，{"；".join(parts)}')
        else:
            print(f'  ✅ {letter} {city_name}：{actual_count} 筆，'
                  f'{letter}00001~{letter}{max_num:05d} 連續，計數器={counter_val}')

    # 5. ID 空白（新增待分配）
    blank_rows = [r for r in rows if not ID_RE.match(str(r.get('ID', '')).strip())]
    if blank_rows:
        print(f'\n  📋 ID 空白（待 assign_ids 分配）：{len(blank_rows)} 筆')
        for r in blank_rows:
            print(f'      → {r.get("店名", "（未知）")} / {r.get("地址", "")}')

    if has_critical:
        print(f'\n  ⚠  偵測到嚴重問題（重複 ID、跳號或計數器不符），流程已暫停，請手動確認。')
    else:
        print(f'\n  ✅ ID 驗證通過')

    return has_critical


def run_path_c():
    print('\n▶ C【完成編輯】正規化 → Excel')

    # ── 1. 記錄舊行數（excel_to_json 前）──────────────────────────────────
    old_row_count = len(load_data())

    # ── 2. Excel → JSON ────────────────────────────────────────────────────
    if os.path.exists(xlsx_path):
        if not step_excel_to_json():
            print('  ❌ Excel 讀取失敗，請檢查後重試。備份（B 路徑建立）保留中。')
            return
    else:
        print('  ℹ  找不到 data.xlsx，直接對 data.json 執行正規化')

    # ── 3. ID 驗證 ─────────────────────────────────────────────────────────
    has_critical = step_validate_ids(old_row_count)

    if has_critical:
        print()
        print('  備份（B 路徑建立）保留中，請手動修正問題後重新執行 C。')
        return

    # ── 4. 正規化流程 ───────────────────────────────────────────────────────
    step_fill_city_district()   # 先補縣市，assign_ids 才能正確判斷城市代碼
    step_assign_ids()
    step_normalize_hours()
    step_normalize_days()
    step_normalize_dates()
    step_auto_close()
    step_sort()
    step_normalize_map_urls(mode='new_only')
    step_json_to_excel()

    # ── 5. 成功 → 刪除 B 路徑建立的備份 ────────────────────────────────────
    _delete_backups()

    print()
    print('═' * 54)
    print('  ✅ 完成！data.json 與 data.xlsx 均已更新')
    print('═' * 54)

while True:
    show_menu()
    choice = input('\n請輸入選項：').strip().lower()

    if choice == 'q':
        print('\n掰掰')
        break

    elif choice == 'a':
        run_path_a()
        input('\n按 Enter 繼續...')

    elif choice == 'b':
        run_path_b()
        input('\n按 Enter 繼續...')

    elif choice == 'c':
        run_path_c()
        input('\n按 Enter 繼續...')

    elif choice == 'd':
        run_path_d()
        input('\n按 Enter 繼續...')

    elif choice == '0':
        while True:
            show_advanced_menu()
            sub = input('\n請輸入數字（b 返回）：').strip().lower()
            if sub == 'b':
                break
            elif sub.isdigit() and 1 <= int(sub) <= len(STEPS):
                _, _, fn = STEPS[int(sub) - 1]
                fn()
                input('\n按 Enter 繼續...')
            else:
                print(f'\n  ⚠  「{sub}」不是有效的選項')

    else:
        print(f'\n  ⚠  「{choice}」不是有效的選項')
        input('\n按 Enter 繼續...')
